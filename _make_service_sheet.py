# -*- coding: utf-8 -*-
"""Builds the service setup sheet the founders fill in for Meevo."""
import sys, os
sys.path.insert(0, "site")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

# menu data, single source of truth lives in site/_build.py
MENU = [
 ("Face", [("Full Face","64"),("Eyebrow","24"),("Cheek","16"),("Chin","14"),
           ("Hairline","14"),("Neck","17"),("Nose","15"),("Sideburn","17"),("Ears","17")]),
 ("Lip", [("Upper Lip","14"),("Lower Lip","14")]),
 ("Bikini", [("Brazilian","64"),("Full Bikini","56"),("Bikini Line","48")]),
 ("Legs", [("Full Leg","79"),("Upper Leg","57"),("Lower Leg","51"),
           ("Inner Thigh","19"),("Knee","17"),("Toes","19")]),
 ("Arms", [("Full Arm","51"),("Half Arm","44"),("Underarm","25"),
           ("Shoulder","31"),("Hand","19")]),
 ("Back", [("Full Back","74"),("Upper Back","31"),("Mid Back","31"),("Lower Back","27")]),
 ("Chest", [("Full Chest","39"),("Chest Strip","27"),("Nipple","17")]),
 ("Stomach", [("Full Stomach","39"),("Stomach Strip","16")]),
 ("Butt", [("Full Butt","33"),("Butt Strip","21")]),
]
WOMEN_ONLY = {"Full Face","Full Butt","Butt Strip","Brazilian","Full Bikini","Bikini Line"}

INK   = "2A2620"; CLAY = "A56B41"; EMERALD = "1D392F"
CREAM = "F5EFE4"; FILLIN = "FFF6D9"; LINE = "DFD5C4"; SOFT = "FBF7EF"

F  = "Arial"
h1 = Font(name=F, size=15, bold=True, color=EMERALD)
h2 = Font(name=F, size=11, bold=True, color="FFFFFF")
cat= Font(name=F, size=10, bold=True, color=CLAY)
bod= Font(name=F, size=10, color=INK)
mut= Font(name=F, size=9,  color="7A7064")
ital=Font(name=F, size=9,  color="7A7064", italic=True)

hdr_fill  = PatternFill("solid", fgColor=EMERALD)
fill_fill = PatternFill("solid", fgColor=FILLIN)
soft_fill = PatternFill("solid", fgColor=SOFT)
cat_fill  = PatternFill("solid", fgColor=CREAM)
thin = Side(style="thin", color=LINE)
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active; ws.title = "Services"

ws["A1"] = "METHOD  ·  Service setup for Meevo online booking"; ws["A1"].font = h1
ws["A2"] = ("Prices are already confirmed from Sophia's email. We need the yellow columns filled in "
            "so the services can be built into Meevo and the Book Now button on the website works.")
ws["A2"].font = mut; ws.merge_cells("A2:H2"); ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 28

ws["A4"] = "Fill in the yellow columns only.  Duration is the one that matters most, Meevo builds the calendar from it."
ws["A4"].font = ital

HEADERS = ["Category","Service","Price","Duration\n(minutes)","Sophia\nperforms?",
           "Tori\nperforms?","Bookable\nonline?","Notes"]
r = 6
for i,h in enumerate(HEADERS, start=1):
    c = ws.cell(row=r, column=i, value=h)
    c.font = h2; c.fill = hdr_fill; c.border = box
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[r].height = 32

ws.cell(row=r, column=4).comment = Comment(
  "How long the appointment takes start to finish, including setup and cleanup.\n"
  "Meevo uses this to build the schedule, so it decides how many guests fit in a day.", "Lacey")

yes_no = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws.add_data_validation(yes_no)

row = r + 1
example_row = None
for category, items in MENU:
    ws.cell(row=row, column=1, value=category).font = cat
    ws.cell(row=row, column=1).fill = cat_fill
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = cat_fill
        ws.cell(row=row, column=col).border = box
    row += 1
    for name, price in items:
        ws.cell(row=row, column=1, value="").border = box
        ws.cell(row=row, column=2, value=name).font = bod
        ws.cell(row=row, column=3, value=int(price)).font = bod
        ws.cell(row=row, column=3).number_format = '"$"#,##0'
        for col in (4,5,6,7):
            c = ws.cell(row=row, column=col); c.fill = fill_fill
        note = "Women only" if name in WOMEN_ONLY else ""
        ws.cell(row=row, column=8, value=note).font = mut
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = box
            if col in (3,4,5,6,7):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        yes_no.add(ws.cell(row=row, column=5))
        yes_no.add(ws.cell(row=row, column=6))
        yes_no.add(ws.cell(row=row, column=7))
        if name == "Brazilian": example_row = row
        row += 1

last = row - 1
ws.cell(row=row+1, column=2, value="Services listed").font = Font(name=F, size=10, bold=True, color=INK)
ws.cell(row=row+1, column=3, value=f'=COUNTA(B{r+1}:B{last})').font = bod
ws.cell(row=row+2, column=2, value="Durations still blank").font = Font(name=F, size=10, bold=True, color=INK)
ws.cell(row=row+2, column=3, value=f'=COUNTA(B{r+1}:B{last})-COUNT(D{r+1}:D{last})').font = bod
ws.cell(row=row+4, column=2,
  value="Example of what we need, using Brazilian: Duration 30, Sophia Yes, Tori Yes, Bookable online Yes.").font = ital

widths = {1:14, 2:20, 3:9, 4:12, 5:11, 6:11, 7:12, 8:24}
for col,w in widths.items(): ws.column_dimensions[get_column_letter(col)].width = w
ws.freeze_panes = f"A{r+1}"

# ---------------- tab 2, details to confirm ----------------
w2 = wb.create_sheet("Please confirm")
w2["A1"] = "Three details do not match, please tell us which is correct"; w2["A1"].font = h1
w2["A2"] = ("Meevo's location record disagrees with the details Sophia emailed. We have built the website on "
            "the email version. Wrong details on a live site also hurt how you rank in local Google searches.")
w2["A2"].font = mut; w2.merge_cells("A2:D2"); w2["A2"].alignment = Alignment(wrap_text=True, vertical="top")
w2.row_dimensions[2].height = 30

for i,h in enumerate(["Detail","What Meevo says","What Sophia's email says","Which is correct?"], start=1):
    c = w2.cell(row=4, column=i, value=h); c.font = h2; c.fill = hdr_fill; c.border = box
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
w2.row_dimensions[4].height = 26

rows = [("Business name","Method Wax Studio","Method Luxury Waxing and Professional Training"),
        ("Phone number","(239) 287-8562","(239) 529-5441"),
        ("ZIP code","Naples, FL 34109","Naples, FL 34108")]
rr = 5
for d,a,b in rows:
    w2.cell(row=rr, column=1, value=d).font = Font(name=F, size=10, bold=True, color=INK)
    w2.cell(row=rr, column=2, value=a).font = bod
    w2.cell(row=rr, column=3, value=b).font = bod
    w2.cell(row=rr, column=4).fill = fill_fill
    for col in range(1,5):
        w2.cell(row=rr, column=col).border = box
        w2.cell(row=rr, column=col).alignment = Alignment(wrap_text=True, vertical="center")
    w2.row_dimensions[rr].height = 30
    rr += 1

rr += 1
w2.cell(row=rr, column=1, value="Two more things when you have a moment").font = Font(name=F, size=11, bold=True, color=EMERALD)
rr += 1
for q in ["Late arrival policy: is it five minutes or ten? We have been told both.",
          "The Year unlimited package: what does it cost?",
          "Parking and how to find Suite 22, a couple of sentences for the website.",
          "Do you have an onboarding contact at Meevo? Setting up services is usually their job and it is included in what you already pay."]:
    w2.cell(row=rr, column=1, value="•  " + q).font = bod
    w2.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=4)
    w2.cell(row=rr, column=1).alignment = Alignment(wrap_text=True, vertical="center")
    w2.row_dimensions[rr].height = 26
    rr += 1

for col,w in {1:26, 2:26, 3:34, 4:26}.items(): w2.column_dimensions[get_column_letter(col)].width = w

out = "METHOD-Service-Setup.xlsx"
wb.save(out); print("saved", out, "| rows:", last-r, "| example row:", example_row)
